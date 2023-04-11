import openpyxl
from openpyxl import load_workbook

# Open the Excel file
workbook = load_workbook(filename="example.xlsx")
sheet = workbook.active

# Select the worksheet you want to work with
worksheet = workbook['تقديري']
# copied_worksheet = workbook.copy_worksheet(worksheet)
# copied_worksheet.title = 'عرض سعر'
# copied_worksheet['A1'] = 'عرض سعر'

# Read data from a cell
index = 0
lastrow = worksheet.max_row
for colm in worksheet.iter_rows(min_row=2 , max_row=2):
    for cell in colm:
        print(cell.value)
        if cell.value == 'السعر الإجمالي':
            index = cell.col_idx
print("the column :",index)
for row in worksheet.iter_rows(min_row=11, max_row=11):
    for cell in row:
        print(cell.value)
# value11 = worksheet[11].value
# print(value11)
print(lastrow)
value = worksheet['A1'].value
print(value)

# Write data to a cell
# worksheet['B1'] = 'Hello, World!'


# Save the changes to the file
# workbook.save('example.xlsx')
